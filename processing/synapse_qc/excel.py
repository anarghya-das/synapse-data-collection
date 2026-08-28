"""Write the quality-results workbook.

Three sheets:
  * ``Summary``     - one row per participant (the headline QC table)
  * ``Per-Channel`` - long format, one row per (participant, channel)
  * ``Legend``      - column definitions, thresholds, and run metadata
"""
import pandas as pd

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAVE_OPENPYXL = False


# Auto-grade thresholds applied to the 0-100 quality_score (% good channels).
GRADE_BANDS = [
    (95, "Excellent"),
    (80, "Good"),
    (60, "Average"),
    (0,  "Bad"),
]

# Fill colours per auto-grade (Summary sheet).
_GRADE_FILL = {
    "Excellent":    "C6EFCE",   # green
    "Good":         "D9EAD3",   # light green
    "Average":      "FFEB9C",   # amber
    "Bad":          "FFC7CE",   # red
    "No EEG":       "D9D9D9",   # grey
    "Empty EEG":    "D9D9D9",
    "Other device": "BDD7EE",   # blue - data exists but wrong headset
    "Error":        "F4B084",   # orange
}

# Map non-"ok" QC statuses to a human grade label.
_STATUS_GRADE = {
    "no_eeg_stream": "No EEG",
    "no_xdf":        "No EEG",
    "empty_eeg":     "Empty EEG",
    "non_obci_eeg":  "Other device",
}


def auto_grade(score, status):
    if status != "ok":
        return _STATUS_GRADE.get(status, "Error")
    for threshold, label in GRADE_BANDS:
        if score >= threshold:
            return label
    return "Bad"


def write_workbook(summary_rows, channel_rows, out_path, run_meta=None):
    """Write the workbook. ``summary_rows`` / ``channel_rows`` are lists of dicts."""
    summary = pd.DataFrame(summary_rows)
    channels = pd.DataFrame(channel_rows)

    # Stable, readable column order for the summary sheet.
    # Manual-rating columns (manual_*) are intentionally absent from the first
    # independent pass; they are appended only if a later comparison step adds
    # them to the rows. Unknown columns are simply skipped below.
    col_order = [
        "participant", "group", "device", "devices_present",
        "auto_grade", "quality_score",
        "score_filtered", "stream_note",
        "n_bad", "n_flat", "n_var", "n_noisy", "n_corr", "n_highcorr",
        "bad_channels", "flat_channels", "dead_channels", "noisy_channels",
        "corr_channels", "highcorr_channels",
        "ear_asymmetry", "mean_snr", "min_snr", "mean_corr",
        "max_corr_bad_frac", "n_corr_windows",
        "n_channels", "duration_s", "sfreq", "method",
        "has_eeg", "has_filtered", "n_bad_filtered",
        "has_responses", "has_video", "n_pdfs",
        "session", "old_only", "neurable", "sub_id", "qc_status",
        "manual_quality", "manual_condition", "manual_severity",
        "manual_notes", "resolve_note",
    ]
    summary = summary[[c for c in col_order if c in summary.columns]]

    legend = _legend_frame(run_meta or {})

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        summary.to_excel(xw, sheet_name="Summary", index=False)
        channels.to_excel(xw, sheet_name="Per-Channel", index=False)
        legend.to_excel(xw, sheet_name="Legend", index=False, header=False)
        if _HAVE_OPENPYXL:
            _style(xw, summary, channels)
    return out_path


def _legend_frame(meta):
    rows = [
        ["SYNAPSE EEG Quality Results", ""],
        ["", ""],
        ["Generated", meta.get("generated", "")],
        ["Preset", meta.get("preset", "")],
        ["Method", meta.get("method", "")],
        ["QC window", meta.get("window", "full recording (no event cropping)")],
        ["Participants analysed", meta.get("n", "")],
        ["", ""],
        ["METHOD = 'robust' (default): all criteria computed on a 1-50 Hz band-passed copy", ""],
        ["", "so RAW (obci_eeg1) and FILTERED (obci_eeg2) inputs converge."],
        ["  flat", "amplitude < flat_voltage for > bad_percent of the time"],
        ["  dead (var)", "SD < sd_floor_uv (absolute low floor)"],
        ["  noisy", "robust z (median/MAD) of log-variance > z_thresh"],
        ["  corr", "MAX off-diagonal correlation < corr_low (disconnected channel)"],
        ["  highcorr", "MAX off-diag corr >= corr_high_report: possible bridging, NOT scored"],
        ["METHOD = 'legacy'", "original metric; criteria on the unfiltered signal (see QC_methodology_review.md)"],
        ["", ""],
        ["AUTO-GRADE (from quality_score = % good channels)", ""],
        ["Excellent", ">= 95"],
        ["Good", "80-94"],
        ["Average", "60-79"],
        ["Bad", "< 60"],
        ["No EEG", "no OpenBCI EEG stream in the XDF"],
        ["", ""],
        ["DUAL STREAMS (raw vs filtered)", ""],
        ["quality_score", "headline = QC on obci_eeg1 (RAW)."],
        ["score_filtered", "same QC on obci_eeg2 (filtered). With the robust method this is a"],
        ["", "CONVERGENCE check: it should ~equal quality_score."],
        ["stream_note", "flags any residual raw-vs-filtered divergence (>=12 pts)"],
        ["", ""],
        ["COLUMN DEFINITIONS", ""],
        ["device", "device of the SCORED recording: OpenBCI (16ch CEEGrid) or Neurable (MW75 14ch)"],
        ["devices_present", "all devices this participant has data for (EXP10 has both, in separate sessions)"],
        ["", "Neurable recordings are NOT scored: the OpenBCI-tuned QC does not apply to a 14ch headset"],
        ["quality_score", "100 * (1 - n_bad / n_channels); n_bad = flat U dead U noisy U corr"],
        ["n_highcorr", "possible bridging; REPORTED, not counted in n_bad"],
        ["ear_asymmetry", "|Lvar - Rvar| / (Lvar + Rvar); 0=symmetric, 1=fully asymmetric"],
        ["mean_snr / min_snr", "per-channel peak-to-peak / MAD, summarised"],
        ["mean_corr", "grand mean of per-channel mean inter-channel correlation (reporting)"],
        ["max_corr_bad_frac", "worst channel's fraction of 1-s windows below corr_low "
                              "(PREP-style windowed criterion; bad if > corr_bad_time_frac)"],
        ["n_corr_windows", "number of 1-s windows the correlation criterion scored"],
        ["session", "sub-folder used; old_only/neurable flag non-standard picks"],
        ["", ""],
        ["Thresholds (preset values)", ""],
    ]
    for k, v in (meta.get("config") or {}).items():
        rows.append([f"  {k}", str(v)])
    return pd.DataFrame(rows)


def _style(xw, summary, channels):
    book = xw.book
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    # ---- Summary sheet ----
    ws = xw.sheets["Summary"]
    grade_col = summary.columns.get_loc("auto_grade") + 1 if "auto_grade" in summary.columns else None
    for j, name in enumerate(summary.columns, start=1):
        c = ws.cell(row=1, column=j)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    # colour each row by auto_grade
    if grade_col:
        for i in range(len(summary)):
            grade = summary.iloc[i]["auto_grade"]
            hexfill = _GRADE_FILL.get(grade)
            if hexfill:
                ws.cell(row=i + 2, column=grade_col).fill = PatternFill("solid", fgColor=hexfill)
    ws.freeze_panes = "C2"
    _autosize(ws, summary)

    # ---- Per-Channel sheet ----
    ws2 = xw.sheets["Per-Channel"]
    for j in range(1, len(channels.columns) + 1):
        c = ws2.cell(row=1, column=j)
        c.font = header_font
        c.fill = header_fill
    ws2.freeze_panes = "A2"
    _autosize(ws2, channels)

    # ---- Legend ----
    ws3 = xw.sheets["Legend"]
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 60
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)


def _autosize(ws, df, cap=40):
    for j, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col].astype(str))) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(width + 2, cap)
